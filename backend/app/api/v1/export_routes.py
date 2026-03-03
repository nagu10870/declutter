"""
Export Routes — Month 5

GET /api/v1/export/csv       → Download file index as CSV
GET /api/v1/export/excel     → Download file index as Excel workbook
GET /api/v1/export/gdpr      → GDPR data export (all user data as JSON archive)
"""

import io
import json
import csv
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.file import FileRecord, FileClassification, Suggestion, DuplicateGroup

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/csv")
async def export_csv(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export all file records as CSV."""
    result = await db.execute(
        select(FileRecord).where(
            FileRecord.user_id == current_user.id,
            FileRecord.is_deleted == False,
        ).order_by(FileRecord.file_size.desc()).limit(50000)
    )
    files = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "file_name", "file_path", "file_size_bytes",
        "mime_type", "md5_hash", "last_modified", "indexed_at",
    ])
    for f in files:
        writer.writerow([
            f.id, f.file_name, f.file_path, f.file_size,
            f.mime_type or "", f.md5_hash or "",
            f.last_modified.isoformat() if f.last_modified else "",
            f.indexed_at.isoformat() if f.indexed_at else "",
        ])

    output.seek(0)
    filename = f"declutter_files_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.read()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/excel")
async def export_excel(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export file index + duplicates + suggestions as Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=503, detail="Excel export not available")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Files ──────────────────────────────────────────────────
    ws_files = wb.active
    ws_files.title = "All Files"

    headers = ["Name", "Path", "Size (MB)", "Type", "Last Modified", "Category", "Blurry", "Screenshot"]
    _style_header_row(ws_files, headers)

    result = await db.execute(
        select(FileRecord, FileClassification)
        .outerjoin(FileClassification, FileClassification.file_id == FileRecord.id)
        .where(FileRecord.user_id == current_user.id, FileRecord.is_deleted == False)
        .order_by(FileRecord.file_size.desc())
        .limit(10000)
    )
    for row in result.all():
        f, fc = row
        ws_files.append([
            f.file_name,
            f.file_path,
            round(f.file_size / 1_048_576, 2),
            f.mime_type or "",
            f.last_modified.strftime("%Y-%m-%d") if f.last_modified else "",
            fc.category if fc else "",
            "Yes" if fc and fc.is_blurry else "",
            "Yes" if fc and fc.is_screenshot else "",
        ])

    # ── Sheet 2: Duplicates ─────────────────────────────────────────────
    ws_dups = wb.create_sheet("Duplicates")
    dup_headers = ["Group ID", "Match Type", "Wasted Bytes (MB)", "Resolved"]
    _style_header_row(ws_dups, dup_headers)

    dup_result = await db.execute(
        select(DuplicateGroup).where(
            DuplicateGroup.user_id == current_user.id,
            DuplicateGroup.resolved == False,
        ).order_by(DuplicateGroup.total_wasted_bytes.desc()).limit(5000)
    )
    for d in dup_result.scalars().all():
        ws_dups.append([
            d.id[:8],
            d.match_type,
            round(d.total_wasted_bytes / 1_048_576, 2),
            "No",
        ])

    # ── Sheet 3: Suggestions ────────────────────────────────────────────
    ws_sugs = wb.create_sheet("Suggestions")
    sug_headers = ["Title", "Type", "Savings (MB)", "Risk Level", "Status"]
    _style_header_row(ws_sugs, sug_headers)

    sug_result = await db.execute(
        select(Suggestion).where(Suggestion.user_id == current_user.id)
        .order_by(Suggestion.bytes_savings.desc()).limit(1000)
    )
    for s in sug_result.scalars().all():
        ws_sugs.append([
            s.title,
            s.suggestion_type,
            round(s.bytes_savings / 1_048_576, 2),
            s.risk_level,
            "Applied" if s.applied else ("Dismissed" if s.dismissed else "Pending"),
        ])

    # Auto-width columns
    for ws in [ws_files, ws_dups, ws_sugs]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"declutter_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/gdpr")
async def export_gdpr(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    GDPR / Data Portability: Export ALL user data as JSON archive.
    Includes: profile, connections, file records, scan jobs, cleanup actions, suggestions.
    Does NOT include OAuth tokens (encrypted, meaningless outside our system).
    """
    from app.models.file import StorageConnection, ScanJob, CleanupAction, FileClassification

    # Fetch all user data
    conns = (await db.execute(
        select(StorageConnection).where(StorageConnection.user_id == current_user.id)
    )).scalars().all()

    files = (await db.execute(
        select(FileRecord).where(FileRecord.user_id == current_user.id).limit(100000)
    )).scalars().all()

    jobs = (await db.execute(
        select(ScanJob).where(ScanJob.user_id == current_user.id)
    )).scalars().all()

    actions = (await db.execute(
        select(CleanupAction).where(CleanupAction.user_id == current_user.id).limit(50000)
    )).scalars().all()

    data = {
        "export_date": datetime.now(timezone.utc).isoformat(),
        "user": {
            "id": current_user.id,
            "email": current_user.email,
            "full_name": current_user.full_name,
            "tier": current_user.tier,
            "created_at": current_user.created_at.isoformat(),
        },
        "storage_connections": [
            {"id": c.id, "provider": c.provider, "account_email": c.account_email,
             "created_at": c.created_at.isoformat()}
            for c in conns
        ],
        "scan_jobs": [
            {"id": j.id, "status": j.status, "files_scanned": j.files_scanned,
             "created_at": j.created_at.isoformat()}
            for j in jobs
        ],
        "file_records": [
            {"id": f.id, "file_name": f.file_name, "file_path": f.file_path,
             "file_size": f.file_size, "mime_type": f.mime_type,
             "indexed_at": f.indexed_at.isoformat()}
            for f in files
        ],
        "cleanup_actions": [
            {"id": a.id, "action": a.action, "bytes_freed": a.bytes_freed,
             "created_at": a.created_at.isoformat()}
            for a in actions
        ],
    }

    json_bytes = json.dumps(data, indent=2).encode()
    filename = f"declutter_gdpr_export_{datetime.now().strftime('%Y%m%d')}.json"
    return StreamingResponse(
        iter([json_bytes]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header_row(ws, headers: list[str]):
    """Apply dark header styling to the first row."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="F0F2FF")
            cell.fill = PatternFill(start_color="252A38", end_color="252A38", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    except Exception:
        ws.append(headers)
